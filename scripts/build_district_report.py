"""
Generate the two district-level download files for /gov/district:

  public/data/downloads/district_report.pdf   one-page district summary
  public/data/downloads/district_report.xlsx  Blocks / Schools / Subjects sheets

Reads src/data/district.json (already renamed to the official unit names) and
the emitted researcher CSVs. English-only (researcher-facing downloads).
Run after build_data.py:  python scripts/build_district_report.py
"""

import csv
import json
import os

import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.normpath(os.path.join(HERE, ".."))
DL = os.path.join(ROOT, "public", "data", "downloads")

district = json.load(open(os.path.join(ROOT, "src", "data", "district.json"), encoding="utf-8"))


def read_csv(name):
    with open(os.path.join(DL, name), encoding="utf-8-sig", newline="") as f:
        return list(csv.reader(f))


# ---------- PDF ----------------------------------------------------------------
GREEN = (14 / 255, 90 / 255, 64 / 255)
INK = (20 / 255, 55 / 255, 38 / 255)
MUTED = (86 / 255, 101 / 255, 121 / 255)

doc = fitz.open()
page = doc.new_page(width=595, height=842)  # A4
y = 56

page.insert_text((56, y), f"{district['name']} District — SAKSHAM Report", fontsize=20, fontname="helv", color=GREEN)
y += 22
page.insert_text((56, y), "Assessment of Grade 5 and Grade 8 students in government schools", fontsize=10, fontname="helv", color=MUTED)
y += 30

stats = [
    ("Overall district score", f"{district['districtAverage']:.0f}%"),
    ("Schools assessed", f"{district['schoolsAssessed']:,}"),
    ("Students assessed", f"{district['studentsAssessed']:,}"),
    ("Blocks", str(len(district["blocks"]))),
    ("Top performing block", district["bestBlock"]),
]
for label, val in stats:
    page.insert_text((56, y), label, fontsize=10, fontname="helv", color=MUTED)
    page.insert_text((220, y), str(val), fontsize=11, fontname="hebo", color=INK)
    y += 17

y += 14
page.insert_text((56, y), "Blocks (sorted by overall score)", fontsize=13, fontname="hebo", color=GREEN)
y += 18
cols = [(56, "Block"), (200, "Overall"), (265, "Grade 5"), (330, "Grade 8"), (395, "Schools"), (460, "Students")]
for x, h in cols:
    page.insert_text((x, y), h, fontsize=9, fontname="hebo", color=MUTED)
y += 6
page.draw_line((56, y), (539, y), color=GREEN, width=0.8)
y += 14
for b in sorted(district["blocks"], key=lambda r: -r["average"]):
    row = [b["name"], f"{b['average']:.1f}%", f"{b['g5']:.1f}%", f"{b['g8']:.1f}%", f"{b['schools']}", f"{b['students']:,}"]
    for (x, _), val in zip(cols, row):
        page.insert_text((x, y), val, fontsize=10, fontname="helv", color=INK)
    y += 16

y += 18
page.insert_text((56, y), "District subject means", fontsize=13, fontname="hebo", color=GREEN)
y += 18
for g in sorted(district["subjectMeans"]):
    means = district["subjectMeans"][g]
    line = "   ".join(f"{s}: {v:.1f}%" for s, v in sorted(means.items()))
    page.insert_text((56, y), f"{g}:  {line}", fontsize=10, fontname="helv", color=INK)
    y += 16

y += 22
page.insert_text((56, y), "All reports are based on the SAKSHAM assessment of Grade 5 and Grade 8 students.", fontsize=8.5, fontname="helv", color=MUTED)

doc.subset_fonts()
doc.save(os.path.join(DL, "district_report.pdf"), deflate=True, garbage=4)
doc.close()
print(f"  wrote district_report.pdf ({os.path.getsize(os.path.join(DL, 'district_report.pdf')) / 1024:.0f} KB)")

# ---------- XLSX ----------------------------------------------------------------
wb = openpyxl.Workbook()


def sheet_from_csv(ws, rows):
    for r in rows:
        ws.append(r)
    for c in range(1, len(rows[0]) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        width = max(len(str(rows[i][c - 1])) for i in range(min(len(rows), 50)))
        ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 10), 42)
    ws.freeze_panes = "A2"


ws = wb.active
ws.title = "Blocks"
sheet_from_csv(ws, read_csv("block_aggregates.csv"))
sheet_from_csv(wb.create_sheet("Schools"), read_csv("schools_overall.csv"))
sheet_from_csv(wb.create_sheet("Subjects by block"), read_csv("block_grade_subject.csv"))

wb.save(os.path.join(DL, "district_report.xlsx"))
print(f"  wrote district_report.xlsx ({os.path.getsize(os.path.join(DL, 'district_report.xlsx')) / 1024:.0f} KB)")
